import pandas as pd
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -- classe de processos --
class AuditProcessor:
    def __init__(self, file_path, et_value=10000, progress_callback=None):
        self.file_path = file_path
        self.et_value = et_value
        self.progress_callback = progress_callback
        self.keywords = ['ajuste', 'estorno', 'erro', 'manual', 'urgente', 'socio']
        
        # - Pré-compilar o Regex uma única vez no início
        self.regex_pattern = re.compile(r'\b(' + '|'.join(self.keywords) + r')\b', re.IGNORECASE)

    # - Começo da Barra de Progresso
    def update_progress(self, value):
        if self.progress_callback:
            self.progress_callback(value)

    def get_objetivo_texto(self, nome_aba):
        objetivos = {
            "Geral": "Apresenta a listagem completa de todos os lançamentos.",
            "ExcedeET": "Filtrar lançamentos acima da materialidade definida.",
            "Redondo": "Identificar lançamentos com valores redondos.",
            "Sem Histórico": "Detectar lançamentos com descrições ausentes.",
            "Final De Semana": "Filtrar lançamentos realizados em sábados ou domingos.",
            "Palavras Chave": "Buscar termos expecíficos no histórico.",
            "Débito x Crédito": "Verificar se os totais de Débito e Crédito estão iguais."
        }
        return objetivos.get(nome_aba.strip(), "Análise contábil.")

    def get_procedimento_texto(self, nome_aba):
        procedimentos = {
            "Geral": "Apenas listagem completa.",
            "ExcedeET": f"Compara Valor_Bruto com ET (R$ {self.et_value}).",
            "Redondo": "Verifica se o valor é maior que zero e múltiplo de 100.",
            "Sem Histórico": "Detecta históricos vazios ou com menos de 2 caracteres.",
            "Final De Semana": "Verifica lançamentos em Sábados (5) e Domingos (6).",
            "Palavras Chave": f"Busca exata pelos termos: {', '.join(self.keywords)}.",
            "Débito x Crédito": "Soma Débito e Crédito para validar equilíbrio."
        }
        return procedimentos.get(nome_aba.strip(), "Procedimento padrão.")

    def get_conclusao_texto(self, nome_aba, stats=None):
        st = stats if stats else {}
        aba = nome_aba.strip()
        if aba == "Final De Semana":
            return f"Detectados {st.get('Fim de Semana', 0)} lançamentos (Sáb: {st.get('sábados', 0)}, Dom: {st.get('domingos', 0)})."
        if aba == "Palavras Chave":
            return f"Encontrados {st.get('Palavras-Chave', 0)} lançamentos críticos: {st.get('detalhe_palavras', '')}."
        if aba == "Débito x Crédito":
            dif = st.get('dif_dc', 0)
            return "Equilíbrio D/C verificado." if dif == 0 else f"Diferença de R$ {dif:,.2f} detectada."
        
        conclusoes = {
            "Geral": "Dados disponibilizados pelo cliente.",
            "ExcedeET": f"Identificados {st.get('Excede ET', 0)} lançamentos acima da materialidade.",
            "Redondo": f"Identificados {st.get('Vlr Redondo', 0)} lançamentos com valores redondos.",
            "Sem Histórico": f"Identificados {st.get('Sem Histórico', 0)} lançamentos com histórico ausente."
        }
        return conclusoes.get(aba, "Procedimento executado conforme planejado.")

    def identificar_palavra(self, texto):
        if pd.isna(texto): return ""
        # - Usa o pattern pré-compilado para ganhar velocidade
        encontradas = self.regex_pattern.findall(str(texto))
        return ", ".join(set(encontradas)).upper()

    def process_audit(self, output_path):
        self.update_progress(0.1)
        
        # - Leitura
        if self.file_path.endswith('.csv'):
            df = pd.read_csv(self.file_path)
        else:
            df = pd.read_excel(self.file_path, engine='calamine')

        df.columns = [str(c).strip() for c in df.columns]
        colunas_originais = list(df.columns)
        
        # - Identificação de colunas
        col_data = next((c for c in df.columns if 'Data' in c), None)
        col_hist = next((c for c in df.columns if 'Histórico' in c or 'Historico' in c), None)
        col_deb = next((c for c in df.columns if 'Débito' in c or 'Debito' in c), None)
        col_cre = next((c for c in df.columns if 'Crédito' in c or 'Credito' in c), None)

        if not all([col_data, col_hist, col_deb, col_cre]):
            raise ValueError("Colunas obrigatórias não encontradas!")

        # - Preparação de Dados
        df['Data_Auxiliar'] = pd.to_datetime(df[col_data], errors='coerce')
        df['Débito'] = pd.to_numeric(df[col_deb], errors='coerce').fillna(0)
        df['Crédito'] = pd.to_numeric(df[col_cre], errors='coerce').fillna(0)
        df['Valor_Bruto'] = df['Débito'] + df['Crédito']
        
        self.update_progress(0.3)

        # - Procedimentos
        df['Excede_ET'] = df['Valor_Bruto'] > self.et_value
        df['Redondo'] = (df['Valor_Bruto'] > 0) & (df['Valor_Bruto'] % 100 == 0)
        df['Sem_Hist'] = (df[col_hist].astype(str).str.len() < 2) | (df[col_hist].isna())
        df['Fds'] = df['Data_Auxiliar'].dt.dayofweek.isin([5, 6])
        df['Palavra_Chave'] = df[col_hist].str.contains(self.regex_pattern, na=False, regex=True)
        df['Termo_Encontrado'] = df[col_hist].apply(self.identificar_palavra)
        
        dif_valor = round(df['Débito'].sum() - df['Crédito'].sum(), 2)
        df['Diferença_Total'] = dif_valor
        df['Status_DC'] = "VALOR IGUAL" if dif_valor == 0 else "VALORES DIFERENTES"
        df['Dia_Da_Semana'] = df['Data_Auxiliar'].dt.dayofweek.map({0:'SEGUNDA',1:'TERÇA',2:'QUARTA',3:'QUINTA',4:'SEXTA',5:'SÁBADO',6:'DOMINGO'})
        
        # - Stats para o relatório
        df_palavras = df[df['Palavra_Chave']]
        contagem_termos = df_palavras['Termo_Encontrado'].str.split(', ').explode().value_counts().to_dict()
        detalhe_str = ", ".join([f"{v} {k}" for k, v in contagem_termos.items()])

        self.stats_detalhada = {
            'Excede ET': int(df['Excede_ET'].sum()),
            'Vlr Redondo': int(df['Redondo'].sum()),
            'Sem Histórico': int(df['Sem_Hist'].sum()),
            'Fim de Semana': int(df['Fds'].sum()),
            'sábados': int((df['Dia_Da_Semana'] == 'SÁBADO').sum()),
            'domingos': int((df['Dia_Da_Semana'] == 'DOMINGO').sum()),
            'Palavras-Chave': int(df['Palavra_Chave'].sum()),
            'detalhe_palavras': detalhe_str.upper(),
            'dif_dc': dif_valor
        }

        # - Escrita no Excel
        procedimentos = {
            "Geral": None, "Débito x Crédito": "FIXO", "ExcedeET": "Excede_ET",
            "Redondo": "Redondo", "Sem Histórico": "Sem_Hist",
            "Final De Semana": "Fds", "Palavras Chave": "Palavra_Chave"
        }

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for idx, (nome_aba, filtro) in enumerate(procedimentos.items()):
                # - Filtragem direta para cada aba, evitando cópias desnecessárias
                if filtro is None or filtro == "FIXO":
                    df_aba = df
                else:
                    df_aba = df[df[filtro] == True].copy()
                    df_aba[filtro] = "VERDADEIRO"

                df_aba[' '] = " "
                
                # - Definição de Colunas por aba
                mapeamento_cols = {
                    "Geral": colunas_originais,
                    "Débito x Crédito": colunas_originais + [' ', 'Diferença_Total', 'Status_DC'],
                    "Palavras Chave": colunas_originais + [' ', 'Palavra_Chave', 'Termo_Encontrado'],
                    "Final De Semana": colunas_originais + [' ', 'Fds', 'Dia_Da_Semana'],
                    "ExcedeET": colunas_originais + [' ', 'Valor_Bruto', 'Excede_ET']
                }
                cols_final = mapeamento_cols.get(nome_aba, colunas_originais + [' ', filtro if filtro else ' '])
                
                df_aba[[c for c in cols_final if c in df_aba.columns]].to_excel(writer, sheet_name=nome_aba, index=False, startrow=10)
                self.aplicar_estilo(writer, nome_aba, len(colunas_originais), (nome_aba != "Geral"))
                self.update_progress(0.4 + (0.6 * (idx + 1) / len(procedimentos)))

        return output_path, self.stats_detalhada

    def aplicar_estilo(self, writer, nome_aba, num_cols_clientes, tem_filtro):
        ws = writer.sheets[nome_aba]
        formato_contabil = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
        fill_cinza = PatternFill("solid", start_color="A6A6A6")
        
        # - Formatação em lote (evita loops pesados em arquivos gigantes)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 4 if col_idx == (num_cols_clientes + 1) else 18
            
            # - Cabeçalho da tabela
            header_cell = ws.cell(row=11, column=col_idx)
            if col_idx != (num_cols_clientes + 1):
                header_cell.fill = fill_cinza
                header_cell.font = Font(bold=True)

            # - Aplica formato contábil se for coluna de valor
            col_name = str(header_cell.value)
            if any(x in col_name for x in ['Débito', 'Crédito', 'Valor', 'Saldo', 'Diferença']):
                for r in range(12, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = formato_contabil

        # - Cabeçalhos Fixos
        ws['A1'] = "@pedro.camposdev Auditoria Ltda."; ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"RELATÓRIO - {nome_aba.upper()}"; ws['A2'].font = Font(size=12, bold=True)
        ws['A3'] = "Cliente:"; ws['A3'].font = Font(size=12, bold=True)
        ws['A4'] = "Objetivo:"; ws['A5'] = self.get_objetivo_texto(nome_aba)
        ws['A6'] = "Procedimento Feito:"; ws['A7'] = self.get_procedimento_texto(nome_aba)
        ws['A8'] = "Conclusão:"; ws['A9'] = self.get_conclusao_texto(nome_aba, getattr(self, 'stats_detalhada', {}))
        for r in [4, 6, 8]: ws[f'A{r}'].font = Font(bold=True)

        # - Marcadores Vermelhos
        ws['A10'] = "x"; ws['A10'].font = Font(color="FF0000", bold=True)
        if tem_filtro:
            y_cell = ws.cell(row=10, column=num_cols_clientes + 2, value="y")
            y_cell.font = Font(color="FF0000", bold=True)
            y_cell.alignment = Alignment(horizontal="center")
        if ws.max_row > 11:
            row_leg = ws.max_row + 2
            ws.cell(row=row_leg, column=1, value="Legenda:").font = Font(name='Arial', size=12, bold=True)
            
            # - Marcador x
            ws.cell(row=row_leg+1, column=1, value="x").font = Font(color="FF0000", bold=True)
            ws.cell(row=row_leg+1, column=2, value="Dados disponibilizados pelo cliente.")
            
            # - Marcador y (apenas se a aba tiver filtro)
            if tem_filtro:
                ws.cell(row=row_leg+2, column=1, value="y").font = Font(color="FF0000", bold=True)
                ws.cell(row=row_leg+2, column=2, value="Dados processados pela auditoria.")
