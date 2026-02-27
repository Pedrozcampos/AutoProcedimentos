import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os
import re
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
import threading

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class AuditProcessor:
    def __init__(self, file_path, et_value=10000, progress_callback=None):
        self.file_path = file_path
        self.et_value = et_value
        self.progress_callback = progress_callback
        self.keywords = ['ajuste', 'estorno', 'erro', 'manual', 'urgente', 'socio']

    def update_progress(self, value):
        if self.progress_callback:
            self.progress_callback(value)

    def get_objetivo_texto(self, nome_aba):
        objetivos = {
            "Geral": "Apresenta a listagem completa de todos os lançamentos.",
            "ExcedeET": "Filtrar lançamentos acima da materialidade definida.",
            "Redondo": "Identificar lançamentos com valores redondos.",
            "Sem Histórico": "Detectar lançamentos com descrições ausentes.",
            "Final De Semana": "Filtrar lançamentos realizados em sábados ou domingos.",
            "Palavras Chave": "Buscar termos expecíficos no histórico.",
            "Débito x Crédito": "Verificar se os totais de Débito e Crédito estão iguais."
        }
        return objetivos.get(nome_aba.strip(), "Análise contábil.")
    def get_procedimento_texto(self, nome_aba):
        procedimento_feito = {
            "Geral": "Apenas listagem completa.",
            "ExcedeET": "O código compara a coluna Valor_Bruto com o ET (o Erro Tolerável ). Se Valor_Bruto > ET, ele marca como positivo.",
            "Redondo": "Ele verifica se o valor é maior que zero e se o resto da divisão por 100 é igual a zero",
            "Sem Histórico": "código marca lançamentos onde o campo 'Histórico' tem menos de 2 caracteres ou está vazio.",
            "Finais De Semana": "Ele converte a coluna 'Data' para o formato datetime e verifica o índice do dia. No Python, 5 é Sábado e 6 é Domingo.",
            "Palavras Chave": "O código varre a coluna ''Histórico'' procurando por: ''ajuste', 'estorno', 'erro', 'manual', 'urgente', 'socio'.",
            "Débito x Crédito": "Soma as colunas de Débito e Crédito para verificar se está equilibrado e se não, mostra o valor de diferença."
        }
        return procedimento_feito.get(nome_aba.strip(), "Procedimento padrão de auditoria.")

    def identificar_palavra(self, texto):
        # - Função auxiliar para retornar qual palavra chave foi encontrada.
        if pd.isna(texto): return ""
        texto = str(texto).lower()
        encontradas = [k for k in self.keywords if k in texto]
        return ", ".join(encontradas).upper()
    
    def process_audit(self, output_path):
        # --- Procedimentos Gerais ---
        self.update_progress(0.08) 

        if self.file_path.endswith('.csv'):
            df = pd.read_csv(self.file_path)
        else:
            df = pd.read_excel(self.file_path, engine='calamine')

        df.columns = [str(c).strip() for c in df.columns]
        colunas_originais = list(df.columns)
        num_cols_clientes = len(colunas_originais)

        # - Identificação Inteligente das Colunas Obrigatórias.
        col_data = next((c for c in df.columns if 'Data' in c), None)
        col_hist = next((c for c in df.columns if 'Histórico' in c or 'Historico' in c), None)
        col_deb = next((c for c in df.columns if 'Débito' in c or 'Debito' in c), None)
        col_cre = next((c for c in df.columns if 'Crédito' in c or 'Credito' in c), None)

        if not all([col_data, col_hist, col_deb, col_cre]):
            raise ValueError(f"Colunas obrigatórias não encontradas! Encontradas: Data({col_data}),"
                            f"Histórico({col_hist}), Débito({col_deb}), Crédito({col_cre})")

        # - Preparação de colunas.
        df['Data_Auxiliar'] = pd.to_datetime(df[col_data], errors='coerce')
        df['Débito'] = pd.to_numeric(df['Débito'], errors='coerce').fillna(0)
        df['Crédito'] = pd.to_numeric(df['Crédito'], errors='coerce').fillna(0)
        df['Valor_Bruto'] = df['Débito'] + df['Crédito']
        total_debito = df['Débito'].sum()
        total_credito = df['Crédito'].sum()

        self.update_progress(0.25)
        
        # - Procedimentos padrão de Auditoria.
        df['Excede_ET'] = df['Valor_Bruto'] > self.et_value
        df['Redondo'] = (df['Valor_Bruto'] > 0) & (df['Valor_Bruto'] % 100 == 0)
        df['Sem_Hist'] = (df['Histórico'].astype(str).str.len() < 2) | (df['Histórico'].isna())
        df['Fds'] = df['Data_Auxiliar'].dt.dayofweek.isin([5, 6])
        df['Palavra_Chave'] = df['Histórico'].str.contains('|'.join(self.keywords), case=False, na=False)
        df['Termo_Encontrado'] = df['Histórico'].apply(self.identificar_palavra)
        df['Diferença_Total'] = self.dif_valor = round(total_debito - total_credito, 2)
        df['Status_DC'] = self.status_eqqui = "Valor igual" if self.dif_valor == 0 else "Valor diferentes"

        # - Lógica de Fim de Semana.
        dias_map = {0: 'Segunda', 1: 'Terça', 2: 'Quarta', 3: 'Quinta', 4: 'Sexta', 5: 'Sábado', 6: 'Domingo'}
        df['Dia_Da_Semana'] = df['Data_Auxiliar'].dt.dayofweek.map(dias_map).str.upper()
        df[col_data] = df['Data_Auxiliar'].dt.date

        # - Formata data para o Excel.
        df['Data'] = df['Data_Auxiliar'].dt.date

        self.update_progress(0.37)
        procedimentos = {
            "Geral": None,
            "Débito x Crédito": "FIXO",
            "ExcedeET": "Excede_ET",
            "Redondo": "Redondo",
            "Sem Histórico": "Sem_Hist",
            "Final De Semana": "Fds",
            "Palavras Chave": "Palavra_Chave"
        }

        # --- Alguns ajustes e definições para o Excel ---
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for idx, (nome_aba, coluna_filtro) in enumerate(procedimentos.items()):
                
                # - Lógica para abas que não filtram linhas (Geral e D/C).
                if coluna_filtro is None or coluna_filtro == "FIXO":
                    df_aba = df.copy()
                else:
                    df_aba = df[df[coluna_filtro] == True].copy()
                
                df_aba[' '] = "" # Coluna de separação visual

                # - Definição de quais colunas aparecem em cada aba.
                if nome_aba == "Geral":
                    cols_final = colunas_originais
                elif nome_aba == "Débito x Crédito":
                    cols_final = colunas_originais + [' ', 'Diferença_Total', 'Status_DC']
                elif nome_aba == "Palavras Chave":
                    df_aba[coluna_filtro] = "VERDADEIRO"
                    cols_final = colunas_originais + [' ', coluna_filtro, 'Termo_Encontrado']
                elif nome_aba == "Final De Semana":
                    df_aba[coluna_filtro] = "VERDADEIRO"
                    cols_final = colunas_originais + [' ', coluna_filtro, 'Dia_Da_Semana']
                elif nome_aba == "ExcedeET":
                    df_aba[coluna_filtro] = "VERDADEIRO"
                    cols_final = colunas_originais + [' ', 'Valor_Bruto', coluna_filtro]
                else:
                    df_aba[coluna_filtro] = "VERDADEIRO"
                    cols_final = colunas_originais + [' ', coluna_filtro]

                # - Garante que apenas colunas existentes no DF sejam exportadas.
                cols_reais = [c for c in cols_final if c in df_aba.columns]
                df_aba[cols_reais].to_excel(writer, sheet_name=nome_aba, index=False, startrow=10)
                
                # - Estilo
                self.aplicar_estilo(writer, nome_aba, num_cols_clientes, (nome_aba != "Geral"))
                self.update_progress(0.4 + (0.5 * (idx + 1) / len(procedimentos)))

        # --- Status final para o dashboard ---
        stats = {
            'Excede ET': int(df['Excede_ET'].sum()),
            'Vlr Redondo': int(df['Redondo'].sum()),
            'Sem Histórico': int(df['Sem_Hist'].sum()),
            'Fim de Semana': int(df['Fds'].sum()),
            'Palavras-Chave': int(df['Palavra_Chave'].sum())
        }
        self.update_progress(1.0)
        return output_path, stats
    
    def aplicar_estilo(self, writer, nome_aba, num_cols_clientes, tem_filtro):
        ws = writer.sheets[nome_aba]
        max_col = ws.max_column
        max_row = ws.max_row

        # - Identificação da coluna de separação (Coluna I, ou seja, a coluna após as colunas do cliente).
        col_sep_idx = num_cols_clientes + 1

        # - Loop de formatação, e ajustes nas linhas e colunas.
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            
            # - Correção na largura das colunas.
            if col_idx == col_sep_idx:
                ws.column_dimensions[col_letter].width = 4
            else:
                ws.column_dimensions[col_letter].width = 18

            for row_idx in range(11, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # - Garante que o texto não quebre (fique cortado).
                if cell.alignment:
                    cell.alignment = Alignment(wrap_text=False, vertical="center", horizontal=cell.alignment.horizontal)
                else:
                    cell.alignment = Alignment(wrap_text=False, vertical="center")

                # - Se for a coluna de separação, Coloca o "espaço" para impedir o texto da 'Conta' de invadir.
                if col_idx == col_sep_idx:
                    if cell.value is None or cell.value == "":
                        cell.value = " "

        # --- CABEÇALHOS (A1-A9) ---
        ws['A1'] = "Empresa @Pedro.camposdev  Ltda."
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws['A2'] = f"RELATÓRIO - {nome_aba.upper()}"
        ws['A2'].font = Font(name='Arial', size=12, bold=True)
        ws['A3'] = f"Cliente:"
        ws['A3'].font = Font(name='Arial', size=12, bold=True)

        ws['A4'] = "Objetivo:"; ws['A4'].font = Font(bold=True)
        ws['A5'] = self.get_objetivo_texto(nome_aba)
        ws['A6'] = "Procedimento Feito:"; ws['A6'].font = Font(bold=True)
        ws['A7'] = self.get_procedimento_texto(nome_aba)
        ws['A8'] = "Conclusão:"; ws['A8'].font = Font(bold=True)

        # - MARCADOR "x".
        ws['A10'] = "x"
        ws['A10'].font = Font(color="FF0000", bold=True)

        # - MARCADOR "y".
        if tem_filtro:
            col_va = num_cols_clientes + 2
            ws.cell(row=10, column=col_va).value = "y"
            ws.cell(row=10, column=col_va).font = Font(color="FF0000", bold=True)
            ws.cell(row=10, column=col_va).alignment = Alignment(horizontal="center")

        # - FORMATAÇÃO DA TABELA (LINHA 11+).
        fill_cinza = PatternFill("solid", start_color="A6A6A6")
        fill_vazio = PatternFill(fill_type=None) # Para limpar a coluna I
        formato_contabil = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
        
        for col_idx in range(1, max_col + 1):
            header = ws.cell(row=11, column=col_idx)
            
            # - PINTA APENAS SE NÃO FOR A COLUNA DE SEPARAÇÃO.
            if col_idx == col_sep_idx:
                header.fill = fill_vazio
            else:
                header.fill = fill_cinza
                header.font = Font(bold=True)
            
            col_name = str(header.value)
            
            # - Formatação contábil.
            if any(x in col_name for x in ['Débito', 'Crédito', 'Valor', 'Saldo']):
                for r in range(12, max_row + 1):
                    cell = ws.cell(row=r, column=col_idx)
                    cell.number_format = formato_contabil

        # - Legenda explicativa de x e y.
        if max_row > 11:
            row_leg = max_row + 2
            ws.cell(row=row_leg, column=1, value="Legenda:").font = Font(name='Arial', size=12, bold=True)
            ws.cell(row=row_leg+1, column=1, value="x").font = Font(color="FF0000", bold=True)
            ws.cell(row=row_leg+1, column=2, value="Dados disponibilizado pelo cliente.   ")
            ws.cell(row=row_leg+2, column=1, value="y").font = Font(color="FF0000", bold=True)
            ws.cell(row=row_leg+2, column=2, value="Dados processados pela auditoria.")

class DashboardWindow(ctk.CTkToplevel):
    # --- Dashboard para mostrar os resultados em gráfico ---
    def __init__(self, stats):
        super().__init__()
        self.title("Painel de Análise")
        self.geometry("800x600")
        
        label = ctk.CTkLabel(self, text="Resumo de Ocorrências por Procedimento", font=("Roboto", 22, "bold"))
        label.pack(pady=20)

        # - Gráfico de Barras.
        fig, ax = plt.subplots(figsize=(8, 5))
        fig.patch.set_facecolor('#1e1e1e')
        ax.set_facecolor('#2b2b2b')
        
        names = list(stats.keys())
        values = list(stats.values())
        
        bars = ax.bar(names, values, color='#1f538d')
        ax.set_xticklabels(names, rotation=45, ha='right', color='white')
        ax.tick_params(axis='y', colors='white')
        
        # - Adiciona rótulos numéricos nas barras.
        for bar in bars:
            yval = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2, yval + 0.1, int(yval), ha='center', color='white', fontweight='bold')

        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=20, pady=20)

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Procedimentos Contábeis - Criado @Pedro.camposdev")
        self.geometry("500x550")
        self.configure(fg_color="#121212") 

        self.main_container = ctk.CTkFrame(self, fg_color="#1e1e1e", corner_radius=25)
        self.main_container.pack(pady=40, padx=40, fill="both", expand=True)

        self.label = ctk.CTkLabel(self.main_container, text="Processamentos Contábeis", font=("Roboto", 27, "bold"))
        self.label.pack(pady=(35, 20))

        self.et_label = ctk.CTkLabel(self.main_container, text="Valor do Erro Tolerável (ET):")
        self.et_label.pack(pady=(10, 5))
        
        self.et_entry = ctk.CTkEntry(self.main_container, width=220, justify="center")
        self.et_entry.insert(0, "10000")
        self.et_entry.pack(pady=10)

        self.btn_run = ctk.CTkButton(self.main_container, text="Selecionar Razão (.xlsx)", command=self.iniciar_processo)
        self.btn_run.pack(pady=25)

        # - Botoes invisíveis inicialmente, só aparecem após o processamento.
        self.progress_bar = ctk.CTkProgressBar(self.main_container, width=300)
        self.progress_bar.set(0)
        self.status_label = ctk.CTkLabel(self.main_container, text="", font=("Arial", 12))
            # -- botao do dashboard -- 
        self.btn_dash = ctk.CTkButton(self.main_container, text="Ver Painel de Análise", command=self.open_dashboard,
                                    fg_color="transparent", border_width=2)
            # -- botao para abrir resultado --
        self.btn_result = ctk.CTkButton(self.main_container, text="Abrir Resultado", command=self.open_result,
                                    fg_color="transparent", border_width=2)

    def atualizar_interface_progresso(self, valor):
        # - Atualiza a barra de progresso com segurança.
        self.progress_bar.set(valor)
        self.status_label.configure(text=f"Progresso: {int(valor*100)}%")

    def iniciar_processo(self):
        path_in = filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.csv")])
        if not path_in: return
        
        path_out = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Razao_Auditado_Final.xlsx")
        if not path_out: return

        self.progress_bar.pack(pady=10)
        self.status_label.pack()
        self.btn_run.configure(state="disabled")

        # - Roda em Thread para não travar a janela.
        thread = threading.Thread(target=self.executar_tarefa, args=(path_in, path_out), daemon=True)
        thread.start()

    def executar_tarefa(self, path_in, path_out):
        try:
            et = float(self.et_entry.get())
            # - Passamos o método da classe App que atualiza a barra como callback.
            proc = AuditProcessor(path_in, et, progress_callback=self.atualizar_interface_progresso)
            _, self.stats = proc.process_audit(path_out)
            self.ultimo_resultado = path_out    

            self.after(0, lambda: self.finalizar_sucesso())
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Erro", f"Falha: {str(e)}"))
            self.after(0, lambda: self.btn_run.configure(state="normal"))
            self.update_progress(0)

    def finalizar_sucesso(self):
        messagebox.showinfo("Sucesso", "Processamento Concluído!")
        self.btn_run.configure(state="normal")
        self.btn_dash.pack(pady=5)
        self.btn_result.pack(pady=5)

    def open_result(self):
        path_saida = self.ultimo_resultado
        if os.path.exists(path_saida):
            os.startfile(path_saida)
        else:
            messagebox.showerror("Erro", "Arquivo de saída não encontrado.")

    def open_dashboard(self):
        if hasattr(self, 'stats'):
            DashboardWindow(self.stats)

if __name__ == "__main__":
    app = App()
    app.mainloop()